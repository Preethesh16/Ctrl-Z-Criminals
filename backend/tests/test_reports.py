"""Exports golden test: full forge case → analyze → all three exports."""

import io
import subprocess
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

FORGE = Path(__file__).resolve().parents[2] / "tools" / "statement-forge"
OUT = FORGE / "out"

client = TestClient(app)


@pytest.fixture(scope="module")
def case_with_analysis():
    subprocess.run([sys.executable, str(FORGE / "forge.py"), str(OUT)], check=True)
    r = client.post("/cases", json={"fir_number": "REPORT-0001/2026", "complainant": "Suresh Patil",
                                    "fraud_amount": "2400000.00"})
    case_id = r.json()["id"]
    for f in sorted(OUT.glob("*")):
        if f.suffix == ".json" or "scanned" in f.name:
            continue
        with f.open("rb") as fh:
            assert client.post(f"/cases/{case_id}/uploads",
                               files={"file": (f.name, fh)}).status_code == 200
    client.post(f"/cases/{case_id}/analyze")
    return case_id


def test_report_preview_html(case_with_analysis):
    r = client.get(f"/cases/{case_with_analysis}/report/preview")
    assert r.status_code == 200
    html = r.text
    assert "Investigation Report" in html
    assert "REPORT-0001/2026" in html
    assert "Round-Trip" in html and "loop-1" in html
    assert "SHA-256" in html or "sha256" in html.lower()


def test_report_pdf(case_with_analysis):
    r = client.get(f"/cases/{case_with_analysis}/export/report.pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 10000


def test_standardized_pdf(case_with_analysis):
    r = client.get(f"/cases/{case_with_analysis}/export/standardized.pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_excel_workbook(case_with_analysis):
    r = client.get(f"/cases/{case_with_analysis}/export/case.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"Transactions", "Flags", "Round Trips",
                                  "Accounts", "Disposition", "Audit"}
    assert wb["Transactions"].max_row > 40
    assert wb["Round Trips"].max_row >= 2  # header + planted loop
    assert wb["Accounts"].max_row >= 9  # 8 statements + header
