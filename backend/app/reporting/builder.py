"""Report data assembly + rendering: HTML (UI preview), PDF (WeasyPrint),
Excel (openpyxl). One data builder feeds every output — preview and court
PDF can never diverge."""

from datetime import UTC, datetime, timedelta
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..models import AnalysisResult, AuditLog, Case, Document, Transaction
from .legal import clauses_for

TEMPLATES = Path(__file__).parent / "templates"
_env = Environment(loader=FileSystemLoader(TEMPLATES), autoescape=select_autoescape(["html"]))

IST = timedelta(hours=5, minutes=30)


def _now_ist() -> str:
    return (datetime.now(UTC) + IST).strftime("%d-%m-%Y %H:%M")


def _artifact(db: Session, case_id: str, kind: str, default):
    row = db.scalar(select(AnalysisResult).where(
        AnalysisResult.case_id == case_id, AnalysisResult.kind == kind))
    return row.payload if row is not None else default


def build_report_context(db: Session, case_id: str) -> dict:
    case = db.get(Case, case_id)
    if case is None:
        raise ValueError("case not found")
    docs = db.scalars(select(Document).where(Document.case_id == case_id)).all()
    txns = db.scalars(select(Transaction).where(Transaction.case_id == case_id)
                      .order_by(Transaction.txn_date, Transaction.row_index)).all()

    documents = []
    for d in docs:
        documents.append({
            "filename": d.filename, "bank_name": d.bank_name, "file_kind": d.file_kind,
            "account_number": d.account_number, "account_holder": d.account_holder,
            "period_from": d.period_from, "period_to": d.period_to, "sha256": d.sha256,
            "txn_count": db.scalar(select(func.count()).select_from(Transaction)
                                   .where(Transaction.document_id == d.id)) or 0,
            "txns": [t for t in txns if t.document_id == d.id],
        })

    flagged = [t for t in txns
               if any(not str(f.get("rule", "")).startswith("_") for f in (t.flags or []))]
    all_rules = {str(f["rule"]) for t in flagged for f in t.flags
                 if not str(f.get("rule", "")).startswith("_")}
    legal = {rule: clauses_for({rule}) for rule in sorted(all_rules) if clauses_for({rule})}

    summary = _artifact(db, case_id, "summary",
                        {"transactions": len(txns),
                         "cleaning": {"duplicate_pairs": 0, "reversal_pairs": 0, "balance_breaks": 0}})
    # summary stores cleaning under its own keys — normalize for template
    cleaning = summary.get("cleaning", summary) if isinstance(summary, dict) else {}
    summary_ctx = {
        "transactions": summary.get("transactions", len(txns)),
        "cleaning": {
            "duplicate_pairs": cleaning.get("duplicate_pairs", 0),
            "reversal_pairs": cleaning.get("reversal_pairs", 0),
            "balance_breaks": cleaning.get("balance_breaks", 0),
        },
    }

    audit = db.scalars(select(AuditLog).where(AuditLog.case_id == case_id)
                       .order_by(AuditLog.at)).all()

    return {
        "css": (TEMPLATES / "base.css").read_text(),
        "generated_at": _now_ist(),
        "case": case,
        "documents": documents,
        "summary": summary_ctx,
        "round_trips": _artifact(db, case_id, "round_trips", []),
        "disposition": _artifact(db, case_id, "disposition",
                                 {"total_debits": "0", "buckets": {}}),
        "correlation": _artifact(db, case_id, "correlation", []),
        "flagged": flagged,
        "legal": legal,
        "audit": [{"at": a.at.strftime("%Y-%m-%d %H:%M:%S") if a.at else "",
                   "actor": a.actor, "action": a.action,
                   "detail": str(a.detail)[:160]} for a in audit],
    }


def render_report_html(db: Session, case_id: str) -> str:
    return _env.get_template("report.html").render(**build_report_context(db, case_id))


def render_standardized_html(db: Session, case_id: str) -> str:
    return _env.get_template("standardized.html").render(**build_report_context(db, case_id))


def html_to_pdf(html: str) -> bytes:
    from weasyprint import HTML

    return HTML(string=html).write_pdf()


def build_excel(db: Session, case_id: str) -> bytes:
    """Multi-sheet workbook: Transactions, Flags, Round Trips, Accounts,
    Disposition, Audit."""
    import io

    from openpyxl import Workbook

    ctx = build_report_context(db, case_id)
    wb = Workbook()

    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Date", "Account", "Narration", "Reference", "Debit", "Credit",
               "Balance", "Channel", "Confidence", "Excluded", "Flags"])
    for d in ctx["documents"]:
        for t in d["txns"]:
            ws.append([
                str(t.txn_date), t.account_ref, t.narration_raw, t.reference_id,
                float(t.amount_inr) if t.direction == "DEBIT" else None,
                float(t.amount_inr) if t.direction == "CREDIT" else None,
                float(t.balance_after) if t.balance_after is not None else None,
                t.channel, t.extraction_confidence, t.excluded,
                "; ".join(str(f.get("rule")) for f in (t.flags or [])),
            ])

    ws = wb.create_sheet("Flags")
    ws.append(["Date", "Account", "Amount", "Direction", "Rule", "Why", "Evidence"])
    for t in ctx["flagged"]:
        for f in t.flags:
            if str(f.get("rule", "")).startswith("_"):
                continue
            extra = {k: v for k, v in f.items() if k not in ("rule", "why")}
            ws.append([str(t.txn_date), t.account_ref, float(t.amount_inr),
                       t.direction, f.get("rule"), f.get("why"), str(extra)])

    ws = wb.create_sheet("Round Trips")
    ws.append(["Loop", "Path", "Hops", "Amount out", "Amount back", "% returned",
               "Elapsed h", "Score"])
    for lp in ctx["round_trips"]:
        ws.append([lp["loop_id"], " -> ".join(lp["path"]), lp["hops"], lp["amount_out"],
                   lp["amount_back"], lp["pct_returned"], lp["elapsed_hours"], lp["score"]])

    ws = wb.create_sheet("Accounts")
    ws.append(["File", "Bank", "Account", "Holder", "Period from", "Period to",
               "Txns", "SHA-256"])
    for d in ctx["documents"]:
        ws.append([d["filename"], d["bank_name"], d["account_number"], d["account_holder"],
                   str(d["period_from"] or ""), str(d["period_to"] or ""),
                   d["txn_count"], d["sha256"]])

    ws = wb.create_sheet("Disposition")
    ws.append(["Bucket", "Amount", "Pct of debits"])
    for name, b in ctx["disposition"].get("buckets", {}).items():
        ws.append([name, b["amount"], b["pct"]])

    ws = wb.create_sheet("Audit")
    ws.append(["When (UTC)", "Actor", "Action", "Detail"])
    for a in ctx["audit"]:
        ws.append([a["at"], a["actor"], a["action"], a["detail"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
